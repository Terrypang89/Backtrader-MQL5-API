import backtrader as bt
import backtrader.indicators as btind
# import backtrader.indicators.MovAv as MovAv
from backtradermql5.mt5store import MTraderStore
from backtradermql5.mt5indicator import getMTraderIndicator
from backtradermql5.mt5chart import MTraderChart, ChartIndicator
from backtradermql5.mt5broker import MTraderBroker
from backtradermql5.mt5data import MTraderData
from datetime import datetime, timedelta, timezone
import time
import json
import pprint as pprint
# from bson import json_util
import pandas as pd
import re
import json
import time
import os
import collections
from flask import Flask, request
from threading import Timer
import csv
import openpyxl
import xlsxwriter
import os.path

# import config
from handler import *

from src import api_server_start
from src import event_subscribe, event_unsubscribe, event_post
from src import StoppableThread
from src import log , create_logger

TIMECALL = 60
API_PORT = "api-port"
API_REV = "api-rev"

class SmaCross(bt.SignalStrategy):
    params = dict(
        smaperiod=5,
        trade=True,
        stake=0.1,
        exectype=bt.Order.Market,
        stopafter=0,
        valid=True,
        cancel=0,
        usebracket=True,
        donotcounter=False,
        sell=True,
    )

    def get_timestamp(self):
        timestamp = time.strftime("%Y-%m-%d %X")
        return timestamp

    def process_data_received(self, data:dict):
        print(f"on_data_received data = {data}\n")
        if not isinstance(data, dict):
            print(f"Incorrect data({data}) format received, SKIP.")
            return None
        try:
            # check if data ticker is 5m, 30m, 1 day
            self._new_data_received = True
            self._api_data = data
            
            return data
        except Exception as err:
            log.error(f"Sent webhook failed, reason: {err}")
        
    def retrieve_data_received(self):
        if self._new_data_received is True:
            self._new_data_received = False
            return self._api_data
        # elif 
        return None
    
    def data_info_print(self):
        txt = list()
        if self.live_data:
            txt.append('LIVE :')
            cash = self.broker.getcash()
        else:
            txt.append('NOT LIVE :')
            cash = "NA"

        txt.append('Data0')
        txt.append('%04d' % len(self.data0))
        dtfmt = '%Y-%m-%dT%H:%M:%S'
        txt.append('{:f}'.format(self.data.datetime[0]))
        txt.append('%s' % self.data.datetime.datetime(0).strftime(dtfmt))
        txt.append('%s' % (self.data.datetime.datetime(0) - timedelta(hours=8)).strftime(dtfmt))
        txt.append('{:f}'.format(self.data.open[0]))
        txt.append('{:f}'.format(self.data.high[0]))
        txt.append('{:f}'.format(self.data.low[0]))
        txt.append('{:f}'.format(self.data.close[0]))
        txt.append('{:6d}'.format(int(self.data.volume[0])))
        txt.append('{:d}'.format(int(self.data.openinterest[0])))
        print(', '.join(txt))
    
    def ind_info_print(self):
        ind_txt = list()
        ind_txt.append('BB1_TOP = {:f}'.format(self.bb1.top[0]))
        ind_txt.append('BB1_BOT = {:f}'.format(self.bb1.bot[0]))
        ind_txt.append('BB1_MID = {:f}'.format(self.bb1.mid[0]))
        ind_txt.append('BB2_TOP = {:f}'.format(self.bb2.top[0]))
        ind_txt.append('BB2_BOT = {:f}'.format(self.bb2.bot[0]))
        ind_txt.append('BB2_MID = {:f}'.format(self.bb2.mid[0]))
        ind_txt.append('BB3_TOP = {:f}'.format(self.bb3.top[0]))
        ind_txt.append('BB3_BOT = {:f}'.format(self.bb3.bot[0]))
        ind_txt.append('BB3_MID = {:f}'.format(self.bb3.mid[0]))

        ind_txt.append('MFI = {:f}'.format(self.mt5mfi.MFI[0]))
        ind_txt.append('rsi_7 = {:f}'.format(self.rsi_7.rsi[0]))
        ind_txt.append('rsi_14 = {:f}'.format(self.rsi_14.rsi[0]))
        
        ind_txt.append('STOCH_MAIN = {:f}'.format(self.mt5stoch.Main[0]))
        ind_txt.append('STOCH_SIGNAL = {:f}'.format(self.mt5stoch.Signal[0]))

        ind_txt.append('MACD = {:f}'.format(self.macd.macd[0]))
        ind_txt.append('MACD_SIGNAL = {:f}'.format(self.macd.signal[0]))
        ind_txt.append('MACD_HISTO = {:f}'.format(self.macd.histo[0]))
        
        ind_txt.append('mt5triema = {:f}'.format(self.mt5triema.TEMA[0]))
        print(', '.join(ind_txt))
    
    def indicator_preset(self, store):

        self.mt5triema = getMTraderIndicator(
            store,
            self.datas[0],
            ("TEMA",),
            indicator="Examples/TEMA",
            params=[3, 0,],
        )()

        self.mt5mfi = getMTraderIndicator(
            store,
            self.datas[0],
            ("MFI",),
            indicator="Examples/MFI",
            params=[7],
        )()

        self.mt5rsi = getMTraderIndicator(
            store,
            self.datas[0],
            ("RSI",),
            indicator="Examples/RSI",
            params=[7,],
        )()

        self.mt5stoch = getMTraderIndicator(
            store,
            self.datas[0],
            ("Main","Signal",),
            indicator="Examples/Stochastic",
            params=[5, 3, 1],
        )()

        self.bb1 = btind.BollingerBands(self.datas[0], period=20, devfactor=1.0, movav=btind.MovAv.Simple)
        self.bb2 = btind.BollingerBands(self.datas[0], period=20, devfactor=2.0, movav=btind.MovAv.Simple)
        self.bb3 = btind.BollingerBands(self.datas[0], period=20, devfactor=3.0, movav=btind.MovAv.Simple)
        self.ema = btind.ExponentialMovingAverage(self.datas[0], period=5)
        self.macd = btind.MACDHisto(self.datas[0], period_me1=12, period_me2=26, period_signal=9, movav=btind.MovAv.Exponential)
        self.rsi_7 = btind.RelativeStrengthIndex(self.datas[0], period=7, movav=btind.MovAv.Smoothed)
        self.rsi_14 = btind.RelativeStrengthIndex(self.datas[0], period=14, movav=btind.MovAv.Smoothed)

        def addChart(chart, ema, bb1, bb2, bb3, macd, rsi_7, rsi_14, mt5mfi, mt5stoch, mt5triema):

            ema_ind1 = ChartIndicator(idx=0, shortname="ema")
            ema_ind1.addline(
                ema.ema,
                style={
                    "linelabel": "EMA",
                    "color": "clrDarkViolet",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            chart.addchartindicator(ema_ind1)

            bb_indi1 = ChartIndicator(idx=0, shortname="Bollinger Bands")
            bb_indi2 = ChartIndicator(idx=0, shortname="Bollinger Bands")
            bb_indi3 = ChartIndicator(idx=0, shortname="Bollinger Bands")

            bb_indi1.addline(
                bb1.top,
                style={
                    "linelabel": "BB1_TOP",
                    "color": "clrBlue",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            bb_indi1.addline(
                bb1.mid,
                style={
                    "linelabel": "BB1_MID",
                    "color": "clrYellow",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            bb_indi1.addline(
                bb1.bot,
                style={
                    "linelabel": "BB1_BOT",
                    "color": "clrBlue",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )

            bb_indi2.addline(
                bb2.top,
                style={
                    "linelabel": "BB2_TOP",
                    "color": "clrCrimson",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            bb_indi2.addline(
                bb2.mid,
                style={
                    "linelabel": "BB2_MID",
                    "color": "clrDarkOrange",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            bb_indi2.addline(
                bb2.bot,
                style={
                    "linelabel": "BB2_BOT",
                    "color": "clrCrimson",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )

            bb_indi3.addline(
                bb3.top,
                style={
                    "linelabel": "BB3_TOP",
                    "color": "clrDarkGreen",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            bb_indi3.addline(
                bb3.mid,
                style={
                    "linelabel": "BB3_MID",
                    "color": "clrDarkOrange",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            bb_indi3.addline(
                bb3.bot,
                style={
                    "linelabel": "BB3_BOT",
                    "color": "clrDarkGreen",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )

            chart.addchartindicator(bb_indi1)
            chart.addchartindicator(bb_indi2)
            chart.addchartindicator(bb_indi3)

            mt5mfi_ind = ChartIndicator(idx=2, shortname="mt5mfi")
            mt5mfi_ind.addline(
                mt5mfi.MFI,
                style={"linelabel": "MFI", "color": "clrLightBlue", "linestyle": "STYLE_SOLID", "linewidth": 2},
            )
            chart.addchartindicator(mt5mfi_ind)

            rsi_ind = ChartIndicator(idx=2, shortname="RSI_7")
            rsi_ind.addline(
                rsi_7.rsi,
                style={"linelabel": "RSI_7", "color": "clrBlue", "linestyle": "STYLE_SOLID", "linewidth": 2},
            )
            chart.addchartindicator(rsi_ind)

            rsi2_ind = ChartIndicator(idx=2, shortname="RSI_14")
            rsi2_ind.addline(
                rsi_14.rsi,
                style={"linelabel": "RSI_14", "color": "clrLightGreen", "linestyle": "STYLE_SOLID", "linewidth": 2},
            )
            chart.addchartindicator(rsi2_ind)

            mt5stoch_ind = ChartIndicator(idx=3, shortname="mt5stoch")
            mt5stoch_ind.addline(
                mt5stoch.Main,
                style={
                    "linelabel": "STOCH_Main",
                    "color": "clrLightGreen",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            mt5stoch_ind.addline(
                mt5stoch.Signal,
                style={
                    "linelabel": "STOCH_Signal",
                    "color": "clrRed",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            chart.addchartindicator(mt5stoch_ind)

            macd1 = ChartIndicator(idx=4, shortname="Macd")
            macd1.addline(
                macd.macd,
                style={"linelabel": "MACD", "color": "clrBlue", "linestyle": "STYLE_SOLID", "linewidth": 2},
            )
            macd1.addline(
                macd.signal,
                style={"linelabel": "MACD_SIGNAL", "color": "clrRed", "linestyle": "STYLE_SOLID", "linewidth": 2},
            )
            macd1.addline(
                macd.histo,
                style={"linelabel": "MACD_HISTO", "color": "clrGreen", "linetype": "DRAW_HISTOGRAM", "linestyle":"STYLE_SOLID", "linewidth": 4},
            )
            chart.addchartindicator(macd1)

            mt5triema_ind = ChartIndicator(idx=0, shortname="mt5triema")
            mt5triema_ind.addline(
                mt5triema.TEMA,
                style={
                    "linelabel": "TEMA",
                    "color": "clrLightGreen",
                    "linestyle": "STYLE_SOLID",
                    "linewidth": 2
                },
            )
            chart.addchartindicator(mt5triema_ind)

        # Instantiate a new chart window and plot
        chart = MTraderChart(self.datas[0], realtime=False)
        addChart(chart, self.ema, self.bb1, self.bb2, self.bb3, self.macd, self.rsi_7, self.rsi_14, self.mt5mfi, self.mt5stoch, self.mt5triema)
        
    def cond1_trend_volatility(self):
        
        # extract the prev_status
        prev_status = None
        prev_result = None
        
        cur_status = None
        cur_result = None
        
        cur_cond1 = {}
        cur_val = {}
        cur_val["bb3"] = {}
        cur_val["bb3"]["top"] = None
        cur_val["bb3"]["bot"] = None
        cur_val["bb2"] = {}
        cur_val["bb2"]["top"] = None
        cur_val["bb2"]["bot"] = None
        cur_val["bb1"] = {}
        cur_val["bb1"]["top"] = None
        cur_val["bb1"]["bot"] = None

        prev_cond1 = {}
        prev_cond1["status"] = {}
        prev_cond1["result"] = {}
        prev_cond1["value"] = {}

        # check if prev status exist 
        if self.cond1_val:
            prev_cond1 = self.cond1_val[-1]
        
        # confirm buy, as bot reversal detected
        if self.mt5triema.TEMA[0] > self.bb3.bot[0] and self.mt5triema.TEMA[-1] < self.bb3.bot[-1]:
            if prev_cond1["status"] == "triema_crossdn_bb3_bot":
                cur_result = "buy_bb3_bot"
            cur_status = "triema_crossup_bb3_bot"
        elif self.mt5triema.TEMA[0] > self.bb2.bot[0] and self.mt5triema.TEMA[-1] < self.bb2.bot[-1]:
            if prev_cond1["status"] == "triema_crossdn_bb2_bot":
                cur_result = "buy_bb2_bot"
            cur_status = "triema_crossup_bb2_bot"
        elif self.mt5triema.TEMA[0] > self.bb1.bot[0] and self.mt5triema.TEMA[-1] < self.bb1.bot[-1]:
            if prev_cond1["status"] == "triema_crossdn_bb1_bot":
                cur_result = "buy_bb1_bot"
            cur_status= "triema_crossup_bb1_bot"

        # set ready buy, triema detect will drossdn bb1, bb2, bb3 low, trade_cond is either 00 = buy entry, 01 = buy exit, 10 = sell entry, 11 = sell exit
        elif self.mt5triema.TEMA[0] < self.bb3.bot[0] and self.mt5triema.TEMA[-1] > self.bb3.bot[-1]:
            if prev_cond1["status"] == "triema_crossdn_bb2_bot":
                cur_result = "ready_buy_bb3_bot"
            cur_status = "triema_crossdn_bb3_bot"
        elif self.mt5triema.TEMA[0] < self.bb2.bot[0] and self.mt5triema.TEMA[-1] > self.bb2.bot[-1]:
            if prev_cond1["status"] == "triema_crossdn_bb1_bot":
                cur_result = "ready_buy_bb2_bot"
            cur_status = "triema_crossdn_bb2_bot"
        elif self.mt5triema.TEMA[0] < self.bb1.bot[0] and self.mt5triema.TEMA[-1] > self.bb1.bot[-1]:
            if prev_cond1["status"] == "triema_crossdn_mid":
                cur_result = "ready_buy_bb1_bot"
            cur_status = "triema_crossdn_bb1_bot"
        elif self.mt5triema.TEMA[0] < self.bb1.mid[0] and self.mt5triema.TEMA[-1] > self.bb1.bot[-1]:
            cur_status = "triema_crossdn_mid"

        # confirm sell, as top reversal detected
        elif self.mt5triema.TEMA[0] < self.bb3.top[0] and self.mt5triema.TEMA[-1] > self.bb3.top[-1]:
            if prev_cond1["status"] == "triema_crossup_bb3_top":
                cur_result = "sell_bb3_top"
            cur_status = "triema_crossdn_bb3_top"
        elif self.mt5triema.TEMA[0] < self.bb2.top[0] and self.mt5triema.TEMA[-1] > self.bb2.top[-1]:
            if prev_cond1["status"] == "triema_crossup_bb2_top":
                cur_result = "sell_bb2_top"
            cur_status = "triema_crossdn_bb2_top"
        elif self.mt5triema.TEMA[0] < self.bb1.top[0] and self.mt5triema.TEMA[-1] > self.bb1.bot[-1]:
            if prev_cond1["status"] == "triema_crossup_bb1_top":
                cur_result = "sell_bb1_top"
            cur_status= "triema_crossdn_bb1_top"

        # ready sell
        elif self.mt5triema.TEMA[0] > self.bb3.top[0] and self.mt5triema.TEMA[-1] < self.bb3.top[-1]:
            if prev_cond1["status"] == "triema_crossup_bb2_top":
                cur_result = "ready_sell_bb3_top"
            cur_status = "triema_crossup_bb3_top"
        elif self.mt5triema.TEMA[0] > self.bb2.top[0] and self.mt5triema.TEMA[-1] < self.bb2.top[-1]:
            if prev_cond1["status"] == "triema_crossup_bb1_top":
                cur_result = "ready_sell_bb2_top"
            cur_status = "triema_crossup_bb2_top"
        elif self.mt5triema.TEMA[0] > self.bb1.top[0] and self.mt5triema.TEMA[-1] < self.bb1.top[-1]:
            if prev_cond1["status"] == "triema_crossup_mid":
                cur_result = "ready_sell_bb1_top"
            cur_status = "triema_crossup_bb1_top"
        elif self.mt5triema.TEMA[0] > self.bb1.mid[0] and self.mt5triema.TEMA[-1] < self.bb1.mid[-1]:
            cur_status = "triema_crossup_mid"
    
        # if cur_status is none while prev_status is not none, set prev_status. if prev_status also none, then set none when no cur_status detected
        if cur_status is None:
            if prev_status is not None:
                cur_status = prev_cond1["status"]
            else:
                cur_status = None

        if cur_result is None:
            if prev_result is not None:
                cur_result = prev_cond1["result"]
            
            else:
                cur_result = None
        
        cur_val["bb3"]["top"] = self.bb3.top[0]
        cur_val["bb3"]["bot"] = self.bb3.bot[0]
        cur_val["bb2"]["top"] = self.bb2.top[0]
        cur_val["bb2"]["bot"] = self.bb2.bot[0]
        cur_val["bb1"]["top"] = self.bb1.top[0]
        cur_val["bb1"]["bot"] = self.bb1.bot[0]

        cur_cond1["value"] = cur_val
        cur_cond1["status"] = cur_status
        cur_cond1["result"] = cur_result
        
        self.cond1_val.append(cur_cond1)

        return self.cond1_val[-1]

    def cond2_momentum(self):

        cur_result = {}
        cur_status = {}
        cur_val = {}

        cur_status["rsi7"] = {}
        cur_status["rsi7"]["value"] = None
        cur_status["rsi7"]["direction"] = None
        cur_status["rsi7"]["trigger"] = None
        cur_status["rsi14"] = {}
        cur_status["rsi14"]["value"] = None
        cur_status["rsi14"]["direction"] = None
        cur_status["rsi14"]["trigger"] = None
        cur_status["stoch"] = {}
        cur_status["stoch"]["value"] = None
        cur_status["stoch"]["direction"] = None
        cur_status["stoch"]["trigger"] = None

        cur_result = {}
        cur_result["rsi7"] = None
        cur_result["rsi14"] = None
        cur_result["stoch"] = None

        cur_val["rsi7"] = {}
        cur_val["rsi14"] = {}
        cur_val["stoch"] = {}

        cur_cond2 = {}

        prev_cond2 = {}
        prev_cond2["status"] = {}
        prev_cond2["value"] = {}
        prev_cond2["result"] = {}
        prev_cond2["status"]["rsi7"] = {}
        prev_cond2["status"]["rsi7"]["value"] = None
        prev_cond2["status"]["rsi7"]["direction"] = None
        prev_cond2["result"]["rsi7"] = None
        prev_cond2["status"]["rsi14"] = {}
        prev_cond2["status"]["rsi14"]["value"] = None
        prev_cond2["status"]["rsi14"]["direction"] = None
        prev_cond2["result"]["rsi14"] = None
        prev_cond2["status"]["stoch"] = {}
        prev_cond2["status"]["stoch"]["value"] = None
        prev_cond2["status"]["stoch"]["direction"] = None
        prev_cond2["result"]["stoch"] = None
        
        # check if prev status exist 
        if self.cond2_val:
            prev_cond2 = self.cond2_val[-1]

        # confirm buy, as bot reversal detected
        if self.rsi_7.rsi[0] > 90 and self.rsi_7.rsi[-1] < 90:
            cur_status["rsi7"]["direction"] = "up"
            cur_status["rsi7"]["value"] = 90
            cur_result["rsi7"] = "ready_buy_exit_" + str(cur_status["rsi7"]["value"])
        elif self.rsi_7.rsi[0] < 90 and self.rsi_7.rsi[-1] > 90:
            cur_status["rsi7"]["direction"] = "dn"
            cur_status["rsi7"]["value"] = 90
            temp_result = "ready_buy_exit_" + str(cur_status["rsi7"]["value"])
            if prev_cond2["result"]["rsi7"] == temp_result: 
                cur_result["rsi7"] = "ready_sell_entry_" + str(cur_status["rsi7"]["value"])
            else:
                cur_result["rsi7"] = prev_cond2["result"]["rsi7"]
        elif self.rsi_7.rsi[0] > 80 and self.rsi_7.rsi[-1] < 80:
            cur_status["rsi7"]["direction"] = "up"
            cur_status["rsi7"]["value"] = 80
            cur_result["rsi7"] = "ready_buy_exit_" + str(cur_status["rsi7"]["value"])
        elif self.rsi_7.rsi[0] < 80 and self.rsi_7.rsi[-1] > 80:
            cur_status["rsi7"]["direction"] = "dm"
            cur_status["rsi7"]["value"] = 80
            temp_result = "ready_buy_exit_" + str(cur_status["rsi7"]["value"])
            if prev_cond2["result"]["rsi7"] == temp_result: 
                cur_result["rsi7"] = "ready_sell_entry_" + str(cur_status["rsi7"]["value"])
            else:
                cur_result["rsi7"] = prev_cond2["result"]["rsi7"]
        elif self.rsi_7.rsi[0] > 70 and self.rsi_7.rsi[-1] < 70:
            cur_status["rsi7"]["direction"] = "up"
            cur_status["rsi7"]["value"] = 70
            cur_result["rsi7"] = "ready_buy_exit_" + str(cur_status["rsi7"]["value"])
        elif self.rsi_7.rsi[0] < 70 and self.rsi_7.rsi[-1] > 70:
            cur_status["rsi7"]["direction"] = "dn"
            cur_status["rsi7"]["value"] = 70
            temp_result = "ready_buy_exit_" + str(cur_status["rsi7"]["value"])
            if prev_cond2["result"]["rsi7"] == temp_result: 
                cur_result["rsi7"] = "ready_sell_entry_" + str(cur_status["rsi7"]["value"])
            else:
                cur_result["rsi7"] = prev_cond2["result"]["rsi7"]

        elif self.rsi_7.rsi[0] > 10 and self.rsi_7.rsi[-1] < 10:
            cur_status["rsi7"]["direction"] = "up"
            cur_status["rsi7"]["value"] = 10
            temp_result = "ready_sell_exit_" + str(cur_status["rsi7"]["value"])
            if prev_cond2["result"]["rsi7"] == temp_result: 
                cur_result["rsi7"] = "ready_buy_entry_" + str(cur_status["rsi7"]["value"])
            else:
                cur_result["rsi7"] = prev_cond2["result"]["rsi7"]
        elif self.rsi_7.rsi[0] < 10 and self.rsi_7.rsi[-1] > 10:
            cur_status["rsi7"]["direction"] = "dn"
            cur_status["rsi7"]["value"] = 10
            cur_result["rsi7"] = "ready_sell_exit_" + str(cur_status["rsi7"]["value"])
        elif self.rsi_7.rsi[0] > 20 and self.rsi_7.rsi[-1] < 20:
            cur_status["rsi7"]["direction"] = "up"
            cur_status["rsi7"]["value"] = 20
            temp_result = "ready_sell_exit_" + str(cur_status["rsi7"]["value"])
            if prev_cond2["result"]["rsi7"] == temp_result: 
                cur_result["rsi7"] = "ready_buy_entry_" + str(cur_status["rsi7"]["value"])
            else:
                cur_result["rsi7"] = prev_cond2["result"]["rsi7"]
        elif self.rsi_7.rsi[0] < 20 and self.rsi_7.rsi[-1] > 20:
            cur_status["rsi7"]["direction"] = "dm"
            cur_status["rsi7"]["value"] = 20
            cur_result["rsi7"] = "ready_sell_exit_" + str(cur_status["rsi7"]["value"])
        elif self.rsi_7.rsi[0] > 30 and self.rsi_7.rsi[-1] < 30:
            cur_status["rsi7"]["direction"] = "up"
            cur_status["rsi7"]["value"] = 30
            temp_result = "ready_sell_exit_" + str(cur_status["rsi7"]["value"])
            if prev_cond2["result"]["rsi7"] == temp_result: 
                cur_result["rsi7"] = "ready_buy_entry_" + str(cur_status["rsi7"]["value"])
            else:
                cur_result["rsi7"] = prev_cond2["result"]["rsi7"]
        elif self.rsi_7.rsi[0] < 30 and self.rsi_7.rsi[-1] > 30:
            cur_status["rsi7"]["direction"] = "dn"
            cur_status["rsi7"]["value"] = 30
            cur_result["rsi7"] = "ready_sell_exit_" + str(cur_status["rsi7"]["value"])
        
        else:
            if self.rsi_7.rsi[0] > 70 or self.rsi_7.rsi[0] < 30:
                cur_status["rsi7"]["direction"] = prev_cond2["status"]["rsi7"]["direction"]
                cur_status["rsi7"]["value"] = prev_cond2["status"]["rsi7"]["value"]
                cur_result["rsi7"] = prev_cond2["result"]["rsi7"]
        # 
        if self.rsi_14.rsi[0] > 90 and self.rsi_14.rsi[-1] < 90:
            cur_status["rsi14"]["direction"] = "up"
            cur_status["rsi14"]["value"] = 90
            cur_result["rsi14"] = "ready_buy_exit_" + str(cur_status["rsi14"]["value"])
        elif self.rsi_14.rsi[0] < 90 and self.rsi_14.rsi[-1] > 90:
            cur_status["rsi14"]["direction"] = "dn"
            cur_status["rsi14"]["value"] = 90
            temp_result = "ready_buy_exit_" + str(cur_status["rsi14"]["value"])
            if prev_cond2["result"]["rsi14"] == temp_result: 
                cur_result["rsi14"] = "ready_sell_entry_" + str(cur_status["rsi14"]["value"])
            else:
                cur_result["rsi14"] = prev_cond2["result"]["rsi14"]
        elif self.rsi_14.rsi[0] > 80 and self.rsi_14.rsi[-1] < 80:
            cur_status["rsi14"]["direction"] = "up"
            cur_status["rsi14"]["value"] = 80
            cur_result["rsi14"] = "ready_buy_exit_" + str(cur_status["rsi14"]["value"])
        elif self.rsi_14.rsi[0] < 80 and self.rsi_14.rsi[-1] > 80:
            cur_status["rsi14"]["direction"] = "dn"
            cur_status["rsi14"]["value"] = 80
            temp_result = "ready_buy_exit_" + str(cur_status["rsi14"]["value"])
            if prev_cond2["result"]["rsi14"] == temp_result: 
                cur_result["rsi14"] = "ready_sell_entry_" + str(cur_status["rsi14"]["value"])
            else:
                cur_result["rsi14"] = prev_cond2["result"]["rsi14"]
        elif self.rsi_14.rsi[0] > 70 and self.rsi_14.rsi[-1] < 70:
            cur_status["rsi14"]["direction"] = "up"
            cur_status["rsi14"]["value"] = 70
            cur_result["rsi14"] = "ready_buy_exit_" + str(cur_status["rsi14"]["value"])
        elif self.rsi_14.rsi[0] < 70 and self.rsi_14.rsi[-1] > 70:
            cur_status["rsi14"]["direction"] = "dn"
            cur_status["rsi14"]["value"] = 70
            temp_result = "ready_buy_exit_" + str(cur_status["rsi14"]["value"])
            if prev_cond2["result"]["rsi14"] == temp_result: 
                cur_result["rsi14"] = "ready_sell_entry_" + str(cur_status["rsi14"]["value"])
            else:
                cur_result["rsi14"] = prev_cond2["result"]["rsi14"]

        elif self.rsi_14.rsi[0] > 10 and self.rsi_14.rsi[-1] < 10:
            cur_status["rsi14"]["direction"] = "up"
            cur_status["rsi14"]["value"] = 10
            temp_result = "ready_sell_exit_" + str(cur_status["rsi14"]["value"])
            if prev_cond2["result"]["rsi14"] == temp_result: 
                cur_result["rsi14"] = "ready_buy_entry_" + str(cur_status["rsi14"]["value"])
            else:
                cur_result["rsi14"] = prev_cond2["result"]["rsi14"]
        elif self.rsi_14.rsi[0] < 10 and self.rsi_14.rsi[-1] > 10:
            cur_status["rsi14"]["direction"] = "dn"
            cur_status["rsi14"]["value"] = 10
            cur_result["rsi14"] = "ready_sell_exit_" + str(cur_status["rsi14"]["value"])
        elif self.rsi_14.rsi[0] > 20 and self.rsi_14.rsi[-1] < 20:
            cur_status["rsi14"]["direction"] = "up"
            cur_status["rsi14"]["value"] = 20
            temp_result = "ready_sell_exit_" + str(cur_status["rsi14"]["value"])
            if prev_cond2["result"]["rsi14"] == temp_result: 
                cur_result["rsi14"] = "ready_buy_entry_" + str(cur_status["rsi14"]["value"])
            else:
                cur_result["rsi14"] = prev_cond2["result"]["rsi14"]
        elif self.rsi_14.rsi[0] < 20 and self.rsi_14.rsi[-1] > 20:
            cur_status["rsi14"]["direction"] = "dn"
            cur_status["rsi14"]["value"] = 20
            cur_result["rsi14"] = "ready_sell_exit_" + str(cur_status["rsi14"]["value"])
        elif self.rsi_14.rsi[0] > 30 and self.rsi_14.rsi[-1] < 30:
            cur_status["rsi14"]["direction"] = "up"
            cur_status["rsi14"]["value"] = 30
            temp_result = "ready_sell_exit_" + str(cur_status["rsi14"]["value"])
            if prev_cond2["result"]["rsi14"] == temp_result: 
                cur_result["rsi14"] = "ready_buy_entry_" + str(cur_status["rsi14"]["value"])
            else:
                cur_result["rsi14"] = prev_cond2["result"]["rsi14"]
        elif self.rsi_14.rsi[0] < 30 and self.rsi_14.rsi[-1] > 30:
            cur_status["rsi14"]["direction"] = "dn"
            cur_status["rsi14"]["value"] = 30
            cur_result["rsi14"] = "ready_sell_exit_" + str(cur_status["rsi14"]["value"])
        else:
            if self.rsi_14.rsi[0] > 70 or self.rsi_14.rsi[0] < 30:
                cur_status["rsi14"]["direction"] = prev_cond2["status"]["rsi14"]["direction"]
                cur_status["rsi14"]["value"] = prev_cond2["status"]["rsi14"]["value"]
                cur_result["rsi14"] = prev_cond2["result"]["rsi14"]

        # check if main crossed signal and above 80
        if self.mt5stoch.Main[0] > self.mt5stoch.Signal[0] and self.mt5stoch.Main[-1] < self.mt5stoch.Signal[-1]:
            if self.mt5stoch.Main[0] >= 90 and self.mt5stoch.Signal[0] >= 90:
                cur_status["stoch"]["value"] = 90
                cur_status["stoch"]["direction"] = "up"
                if cur_status["rsi7"]["trigger"]: 
                    if "ready_buy_exit" in cur_status["rsi7"]["trigger"] or "ready_sell_entry" in cur_status["rsi7"]["trigger"]:
                        cur_status["stoch"]["trigger"] = "strong_sell"

            elif self.mt5stoch.Main[0] >= 80 and self.mt5stoch.Signal[0] >= 80:
                cur_status["stoch"]["value"] = 80
                cur_status["stoch"]["direction"] = "up"
                if cur_status["rsi7"]["trigger"]:
                    if "ready_buy_exit" in cur_status["rsi7"]["trigger"] or "ready_sell_entry" in cur_status["rsi7"]["trigger"]:
                        cur_status["stoch"]["trigger"] = "sell"

            # elif self.mt5stoch.Main[0] > 70 and self.mt5stoch.Signal[0] > 70:
            #     cur_status["stoch"]["value"] = 70
            #     cur_status["stoch"]["direction"] = "up"
            #     if cur_status["rsi7"]["trigger"]:
            #         if "ready_buy_exit" in cur_status["rsi7"]["trigger"] or "ready_sell_entry" in cur_status["rsi7"]["trigger"]:
            #             cur_status["stoch"]["trigger"] = "strong_sell"

            # else:
            #     if self.mt5stoch.Main[0] > 70 and self.mt5stoch.Signal[0] > 70:
            #         cur_status["stoch"]["value"] = prev_cond2["status"]["stoch"]["value"]
            #         cur_status["stoch"]["direction"] = prev_cond2["status"]["stoch"]["direction"]
                    # if "ready_buy_exit" in cur_status["rsi14"]["trigger"] or "ready_sell_entry" in cur_status["rsi14"]["trigger"]:
                        # cur_status["stoch"]["trigger"] = "ready_sell"

        elif self.mt5stoch.Main[0] < self.mt5stoch.Signal[0] and self.mt5stoch.Main[-1] > self.mt5stoch.Signal[-1]:
            # if self.mt5stoch.Main[0] < 30 and self.mt5stoch.Signal[0] < 30:
            #     cur_status["stoch"]["value"] = 30
            #     cur_status["stoch"]["direction"] = "dn"
            #     if cur_status["rsi7"]["trigger"]:
            #         if "ready_sell_exit" in cur_status["rsi7"]["trigger"] or "ready_buy_entry" in cur_status["rsi7"]["trigger"]:
            #             cur_status["stoch"]["trigger"] = "buy"

            if self.mt5stoch.Main[0] <= 20 and self.mt5stoch.Signal[0] <= 20:
                cur_status["stoch"]["value"] = 20
                cur_status["stoch"]["direction"] = "dn"
                if cur_status["rsi7"]["trigger"]:
                    if "ready_sell_exit" in cur_status["rsi7"]["trigger"] or "ready_buy_entry" in cur_status["rsi7"]["trigger"]:
                        cur_status["stoch"]["trigger"] = "buy"
            elif self.mt5stoch.Main[0] <= 10 and self.mt5stoch.Signal[0] <= 10:
                cur_status["stoch"]["value"] = 10
                cur_status["stoch"]["direction"] = "dn"
                if cur_status["rsi7"]["trigger"]:
                    if "ready_sell_exit" in cur_result["rsi7"] or "ready_buy_entry" in cur_result["rsi7"]:
                        cur_result["stoch"] = "strong_buy"
            # else:
            #     if self.mt5stoch.Main[0] < 30 and self.mt5stoch.Signal[0] < 30:
            #         cur_status["stoch"]["value"] = prev_cond2["status"]["stoch"]["value"]
            #         cur_status["stoch"]["direction"] = prev_cond2["status"]["stoch"]["direction"]


        cur_val["rsi7"] = self.rsi_7.rsi[0]
        cur_val["rsi14"] = self.rsi_14.rsi[0]
        cur_val["stoch"]["main"] = self.mt5stoch.Main[0]
        cur_val["stoch"]["signal"] = self.mt5stoch.Signal[0]

        cur_cond2["value"] = cur_val
        cur_cond2["status"] = cur_status
        cur_cond2["result"] = cur_result

        self.cond2_val.append(cur_cond2)

        return self.cond2_val[-1]
    
    def cond3_divergence(self):
        
        cur_hist = {}
        cur_hist["status"] = None
        cur_hist["direction"] = None
        cur_hist["num"] = None
        cur_hist["times"] = None
        cur_hist["stage"] = None
        cur_hist["trend"] = None

        cur_macd = {}
        cur_macd["direction"] = None
        cur_macd["trend"] = None
        cur_macd["status"] = None

        cur_signal = {}
        cur_signal["status"] = None
        cur_signal["trend"] = None
        cur_signal["status"] = None

        cur_status = {}
        cur_val = {}
    
        cur_cond3 = {}

        prev_cond3 = {}
        prev_cond3["status"] = {}
        prev_cond3["status"]["macd"] = {}
        prev_cond3["status"]["signal"] = {}
        prev_cond3["status"]["hist"] = {}
        prev_cond3["status"]["macd"]["trend"] = None
        prev_cond3["status"]["signal"]["trend"] = None
        prev_cond3["status"]["macd"]["prev_crossed_val"] = None
        prev_cond3["status"]["signal"]["prev_crossed_val"] = None
        prev_cond3["status"]["hist"]["num"] = None
        prev_cond3["status"]["hist"]["times"] = None
        prev_cond3["status"]["hist"]["direction"] = None
        prev_cond3["status"]["hist"]["trend"] = None
        prev_cond3["status"]["hist"]["status"] = None

        cur_hist["trigger"] = None

        # check if prev array exist
        if self.cond3_val:
            prev_cond3 = self.cond3_val[-1]

        # check macd and signal at top or bottom of 0
        if self.macd.macd[0] > 0:
            cur_macd["direction"] = "up"
        else:
            cur_macd["direction"] = "dn"
        if self.macd.signal[0] > 0:
            cur_signal["direction"] = "up"
        else:
            cur_signal["direction"] = "dn"

        # macd crossed signal 
        if self.macd.macd[0] > self.macd.signal[0] and self.macd.macd[-1] < self.macd.signal[-1]:
            # macd crossed signal above 0, below 0
            if cur_macd["direction"] == "up" and cur_signal["direction"] == "up":
                cur_macd["trend"] = "dn>0"
                cur_signal["trend"] = "dn>0"
            elif cur_macd["direction"] == "dn" and cur_signal["direction"] == "dn":
                cur_macd["trend"] = "dn<0"
                cur_signal["trend"] = "dn<0"
            else:
                cur_macd["trend"] = "dn=0"
                cur_signal["trend"] = "dn=0"
            cur_hist["trigger"] = "macd_crossed_signal"
            cur_macd["prev_crossed_val"] = self.macd.macd[0]
            cur_signal["prev_crossed_val"] = self.macd.signal[0]

        # signal crossed macd
        elif self.macd.macd[0] < self.macd.signal[0] and self.macd.macd[-1] > self.macd.signal[-1]:
            # signal crossed macd above 0, below 0
            if cur_macd["direction"] == "up" and cur_signal["direction"] == "up":
                cur_macd["trend"] = "up>0"
                cur_signal["trend"] = "up>0"
            elif cur_macd["direction"] == "dn" and cur_signal["direction"] == "dn":
                cur_macd["trend"] = "up<0"
                cur_signal["trend"] = "up<0"
            else:
                cur_macd["trend"] = "up=0"
                cur_signal["trend"] = "up=0"
            cur_hist["trigger"] = "signal_crossed_macd"
            cur_macd["prev_crossed_val"] = self.macd.macd[0]
            cur_signal["prev_crossed_val"] = self.macd.signal[0]
        else:
            if prev_cond3["status"]["macd"]["trend"] and prev_cond3["status"]["macd"]["prev_crossed_val"]:
                # when no crossed, check prev macd trend is up or dn,
                if "up" in prev_cond3["status"]["macd"]["trend"]:
                    #check cur macd val is below or above prev macd crossed value
                    if self.macd.macd[0] < prev_cond3["status"]["macd"]["prev_crossed_val"] and self.macd.signal[0] < prev_cond3["status"]["signal"]["prev_crossed_val"]:
                        cur_macd["trend"] = "up"
                        cur_signal["trend"] = "up"
                elif "dn" in prev_cond3["status"]["macd"]["trend"]:
                    if self.macd.macd[0] > prev_cond3["status"]["macd"]["prev_crossed_val"] and self.macd.signal[0] > prev_cond3["status"]["signal"]["prev_crossed_val"]:
                        cur_macd["trend"] = "dn"
                        cur_signal["trend"] = "dn"
            else:
                cur_macd["trend"] = None
                cur_signal["trend"] = None
            cur_macd["prev_crossed_val"] = prev_cond3["status"]["macd"]["prev_crossed_val"]
            cur_signal["prev_crossed_val"] = prev_cond3["status"]["signal"]["prev_crossed_val"]

        # check cur hist num = prev status 
        if not prev_cond3["status"]["hist"]["num"]:
            cur_hist["num"] = 0
        else:
            cur_hist["num"] = prev_cond3["status"]["hist"]["num"]
        
        if not prev_cond3["status"]["hist"]["times"]:
            cur_hist["times"] = 0
        else:
            cur_hist["times"] = prev_cond3["status"]["hist"]["times"]
            

        # detect hist crossed 0, 
        if self.macd.histo[0] > 0 and self.macd.histo[-1] < 0:
            cur_hist["direction"] = "up>0"
            # cur_hist["status"] = "new_up"
            cur_hist["status"] = "uptrend"
            cur_hist["num"] = cur_hist["num"] + 1
            cur_hist["times"] = 0
        elif self.macd.histo[0] < 0 and self.macd.histo[-1] > 0:
            cur_hist["direction"] = "dn<0"
            # cur_hist["status"] = "new_dn"
            cur_hist["status"] = "dntrend"
            cur_hist["num"] = cur_hist["num"] + 1
            cur_hist["times"] = 0
        elif self.macd.histo[0] > 0 and self.macd.histo[-1] > 0:
            cur_hist["direction"] = "up"
        elif self.macd.histo[0] < 0 and self.macd.histo[-1] < 0:
            cur_hist["direction"] = "dn"

        # detect hist up, then hist peak, hist drop
        if cur_hist["direction"]:
            if "up" in cur_hist["direction"]:
                if self.macd.histo[0] < self.macd.histo[-1] and self.macd.histo[-1] > self.macd.histo[-2] and self.macd.histo[-2] > self.macd.histo[-3]:
                    # cur_hist["trigger"] = "hist_num_peak"
                    cur_hist["status"] = "peak"
                elif self.macd.histo[0] > self.macd.histo[-1] and self.macd.histo[-1] < self.macd.histo[-2] and self.macd.histo[-2] < self.macd.histo[-3]:
                    # cur_hist["trigger"] = "hist_num_drop"
                    cur_hist["status"] = "drop"
                    cur_hist["times"] = cur_hist["times"] + 1
                
                # when no status, if prev status is peak, cur status should be dntrend, elif prev status is drop, cur status should be uptrend, else inherit from prev status
                if not cur_hist["status"]:
                    if prev_cond3["status"]["hist"]["status"] == "peak":
                        if self.macd.histo[0] < self.macd.histo[-1] and self.macd.histo[-1] < self.macd.histo[-2]:
                            cur_hist["status"] = "dntrend"

                    elif prev_cond3["status"]["hist"]["status"] == "drop":
                        if self.macd.histo[0] > self.macd.histo[-1] and self.macd.histo[-1] > self.macd.histo[-2]:
                            cur_hist["status"] = "dntrend"
                    else:
                        cur_hist["status"] = prev_cond3["status"]["hist"]["status"]

            elif "dn" in cur_hist["direction"]:
                if self.macd.histo[0] < self.macd.histo[-1] and self.macd.histo[-1] > self.macd.histo[-2] and self.macd.histo[-2] > self.macd.histo[-3]:
                    cur_hist["status"] = "drop"
                    cur_hist["times"] = cur_hist["times"] + 1
                elif self.macd.histo[0] > self.macd.histo[-1] and self.macd.histo[-1] < self.macd.histo[-2] and self.macd.histo[-2] < self.macd.histo[-3]:
                    cur_hist["status"] = "peak"

                if not cur_hist["status"]:
                    if prev_cond3["status"]["hist"]["status"] == "peak":
                        if self.macd.histo[0] > self.macd.histo[-1] and self.macd.histo[-1] > self.macd.histo[-2]:
                            cur_hist["status"] = "uptrend"
                    
                    elif prev_cond3["status"]["hist"]["status"] == "drop":
                        if self.macd.histo[0] < self.macd.histo[-1] and self.macd.histo[-1] < self.macd.histo[-2]:
                            cur_hist["status"] = "dntrend"
                    else:
                        cur_hist["status"] = prev_cond3["status"]["hist"]["status"]

        # set a none word for concatenate whole words
        cur_hist_status = cur_hist["status"]
        if not cur_hist["status"]:
            cur_hist_status = "None"

        cur_hist["trigger"] = str(cur_hist["num"]) + "_" + cur_hist["direction"] + "_" + cur_hist_status + "_" + str(cur_hist["times"])

        # store value
        cur_status["macd"] = cur_macd
        cur_status["signal"] = cur_signal
        cur_status["hist"] = cur_hist

        cur_val["macd"] = self.macd.macd[0]
        cur_val["signal"] = self.macd.signal[0]
        cur_val["hist"] = self.macd.histo[0]

        cur_cond3["value"] = cur_val
        cur_cond3["status"] = cur_status

        self.cond3_val.append(cur_cond3)

        return self.cond3_val[-1]
    
    def entry_with_no_trades(self):
        # cond1_val = self.cond1_trend_volatility(0)
        return None
    
    def entry_with_buy_sell_trades(self):
        # check prev trade 
        # prev_trade = self.orders[self.orderid-1][]
        # cond1_val = self.cond1_trend_volatility(0)
        return None
    
    def exit_buy_sell_trades(self):

        # checck close trade, then orderid + 1 

        # make sure trade is open
        # if self.orderid == 0:
        #     if self.cond1_trend_volatility() != 0 and self.cond2_momentum() != 0:

        # else:
        #     for extracted_order in self.orders:
        #         # if updated_data["Alert"] == "SUPER_BUY" or updated_data["Alert"] == "SUPER_SELL":
        #         #     print(f"CANCELING order {extracted_order} with alert {updated_data['Alert']}")
        #             self.cancel(extracted_order)
        #     self.orders[self.orderid]["trade_cond"] = 0
        # self.orders[self.orderid]["exit_reason"] = ""
        # self.orderid += 1
        #     print(f"CHECKING orderid {self.orderid}, datastatus = {self.datastatus}, position = {self.position}")
        #     print("=============================================")

        return None
    
    def __init__(self, store):

        self.orderid = 0
        self.orders = None
        self.live_data = None

        self.counttostop = 0
        self.datastatus = 0

        self._new_data_received = False
        self._api_data = ""

        self.cond1_val = collections.deque([], maxlen=50)
        self.cond2_val = collections.deque([], maxlen=50)
        self.cond3_val = collections.deque([], maxlen=50)

        self.cond_table = collections.deque([], maxlen=50)

        # self.buy_orders = list()
        # self.sell_orders = list()
        
        self.cur_trade = None
        self.prev_trade = None
        self.cur_trend = None
        self.prev_trade = None
        self.trade_cond = None

        event_subscribe(API_REV, self.process_data_received)

        self.indicator_preset(store)

    def obj2array(self, temp_name, temp_array, obj):
        prev_name = temp_name
        for i,j in obj.items():
            if not isinstance(j, dict):
                # print(f"i = {i}, temp_name = {temp_name}")
                temp_array[temp_name + "_" + str(i)] = j
            else:
                temp_name = prev_name + "_" + i
                self.obj2array(temp_name, temp_array, j) 

    def write_excel(self, filename, sheetname, dataframe):
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            workBook = writer.book
            try:
                print(f"sheet {sheetname} detected, replacing!")
                workBook.remove(workBook[sheetname])
            except:
                print("Worksheet does not exist")
            finally:
                dataframe.to_excel(writer, sheet_name=sheetname,index=False)
                writer.save()

    def next(self):

        # data and indicator printing
        self.data_info_print()
        # self.ind_info_print()
        
        if not self.p.trade:
            return
        
        if not self.live_data:
            return

        cond1 = self.cond1_trend_volatility()
        print(f"cond1 = {json.dumps(cond1, indent=4)}, \n")

        cond2 = self.cond2_momentum()
        print(f"cond2 = {json.dumps(cond2, indent=4)}, \n")

        cond3 = self.cond3_divergence()
        print(f"cond3 = {json.dumps(cond3, indent=4)}, \n")

        self.temp_cond = {}
        self.temp_name = "cond1"

        dtfmt = '%Y-%m-%dT%H:%M:%S'
        self.temp_cond["time"] = str((self.data.datetime.datetime(0) - timedelta(hours=8)).strftime(dtfmt))
        self.obj2array(self.temp_name, self.temp_cond, cond1)
        # print(f"temp_cond1 = {self.temp_cond1}\n")

        # self.temp_cond2 = {}
        self.temp_name = "cond2"
        self.obj2array(self.temp_name, self.temp_cond, cond2)
        # print(f"temp_cond2 = {self.temp_cond2}\n")
                
        # self.temp_cond3 = {}
        self.temp_name = "cond3"
        self.obj2array(self.temp_name, self.temp_cond, cond3)
        # print(f"temp_cond = {self.temp_cond}")

        # self.cond_table.append(self.temp_cond)
        # store all to panda dataframe, 
        # json_array = pd.DataFrame(self.cond_table)
        # sheet_time = datetime.datetime.now()
        # self.write_excel("test.xlsx", "hello", json_array)
        # print(json_array.to_string())

        index_list = list(self.temp_cond.keys())
        value_list = list(self.temp_cond.values())

        # python check if xlsx file exist then use load_workbook
        if os.path.isfile("output.xlsx"):
            wb = openpyxl.load_workbook("output.xlsx")
            ws = wb.active
            ws.append(value_list)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(index_list)
            ws.append(value_list)
        wb.save('output.xlsx')

        updated_data = self.retrieve_data_received()
        # if no new api data received then exit
        if updated_data is None:
            return
        
        print(f"UPDATED data = {updated_data}")
        print(f"PRECHECK datastatus = {self.datastatus}, position = {self.position}, orderid = {self.orderid}, orders = {self.orders}, usebracket = {self.p.usebracket}, donorcounter = {self.p.donotcounter}, valid = {self.p.valid}, stake = {self.p.stake}, exectype={self.p.exectype}")

        # if self.orders[self.tradeid] is  None:
        #     if self.buy_orders is None and self.sell_orders is None:
        #         self.entry_with_no_trades()
        #     else:
        #         self.entry_with_buy_sell_trades()
        #     self.trade_cond = "open"

        # elif self.trade_cond == "open":
        #     self.exit_buy_sell_trades()

        self.orders = None
        # check if ordernum is 
        for ordernum in range(0,1):
            if self.orders[self.orderid]["order"][ordernum] is not None:
                if self.orders[self.orderid]["order_type"] == ordernum:
                    if updated_data["Alert"] == "SUPER_BUY" or updated_data["Alert"] == "SUPER_SELL":
                        self.cancel(self.orders[self.orderid]["order"][ordernum])

        # # if self.datastatus and need make sure orderid is not 0, so we can detect prev trade condition
        if self.datastatus and self.orders[self.orderid]["buy"] is None and self.orders[self.orderid]["sell"] is None:
            if not self.p.usebracket:
                print('USING WITHOUT BRACKET')
                if updated_data["Alert"] == "SUPER_BUY":
                    # check order
                    price = self.data0.close[0]
                    print(f"BUY ACTION on price {price}")
                    self.order = self.buy(size=self.p.stake,
                                          exectype=self.p.exectype,
                                          price=price,
                                          valid=self.p.valid)
                    self.orders[self.orderid]["order_exetype"] = self.p.exectype
                    self.orders[self.orderid]["order_type"] = "buy"
                    self.orders[self.orderid]["order"]["buy"].append(self.order)
                    # self.orders[self.orderid]["trade_cond"] = 1
                elif updated_data["Alert"] == "SUPER_SELL":
                    # price = round(self.data0.close[0] * 1.10, 4)
                    price = self.data0.close[0]
                    print(f"SELL ACTION on price {price}")
                    self.order = self.sell(size=self.p.stake,
                                           exectype=self.p.exectype,
                                           price=price,
                                           valid=self.p.valid)
                    self.orders[self.orderid]["order_exetype"] = self.p.exectype
                    self.orders[self.orderid]["order_type"] = "sell"
                    self.orders[self.orderid]["order"]["sell"].append(self.order)

            else:
                if updated_data["Alert"] == "SUPER_BUY":
                    
                    price = self.data0.close[0]
                    stopprice=price - 10.00
                    limitprice=price + 10.00
                    print(f'USING BUY BRACKET with price {price}, stopprice {stopprice}, limit price {limitprice}')
                    self.order, _, _ = self.buy_bracket(size=self.p.stake,
                                                        exectype=bt.Order.Market,
                                                        price=price,
                                                        stopprice=stopprice,
                                                        limitprice=limitprice,
                                                        valid=self.p.valid)
                    self.orders[self.orderid]["order_exetype"] = "bracket"
                    self.orders[self.orderid]["order_type"] = "buy"
                    self.orders[self.orderid]["order"]["buy"].append(self.order)
                elif updated_data["Alert"] == "SUPER_SELL":
                    price = self.data0.close[0]
                    stopprice=price + 10.00
                    limitprice=price - 10.00
                    print(f'USING SELL BRACKET with price {price}, stopprice {stopprice}, limit price {limitprice}')
                    self.order, _, _ = self.sell_bracket(size=self.p.stake,
                                                        exectype=bt.Order.Market,
                                                        price=price,
                                                        stopprice=stopprice,
                                                        limitprice=limitprice,
                                                        valid=self.p.valid)
                    self.orders[self.orderid]["order_exetype"] = "bracket"
                    self.orders[self.orderid]["order_type"] = "sell"
                    self.orders[self.orderid]["order"]["sell"].append(self.order)


            # self.orders.append(self.order)
        elif self.position and not self.p.donotcounter:
            if self.orders[self.orderid]["buy"] is None and self.orders[self.orderid]["sell"] is None:
                if updated_data["Alert"] == "SUPER_SELL":
                    print(f"SELL with POSITION at price {self.data0.close[0]}")
                    self.order = self.sell(size=self.p.stake // 2,
                                           exectype=bt.Order.Market,
                                           price=self.data0.close[0])
                    self.orders[self.orderid]["order_exetype"] = "market"
                    self.orders[self.orderid]["order_type"] = "sell"
                    self.orders[self.orderid]["sell"].append(self.order)
                elif updated_data["Alert"] == "SUPER_BUY":
                    print(f"BUY with POSITION at price {self.data0.close[0]}")
                    self.order = self.buy(size=self.p.stake // 2,
                                          exectype=bt.Order.Market,
                                          price=self.data0.close[0])
                    self.orders[self.orderid]["order_exetype"] = "market"
                    self.orders[self.orderid]["order_type"] = "sell"
                    self.orders[self.orderid]["buy"].append(self.order)


        if self.datastatus:
            self.datastatus += 1


    def notify_store(self, msg, *args, **kwargs):
        print('*' * 5, 'STORE NOTIF notify_store:', msg)

    def notify_order(self, order):
        if order.status in [order.Completed, order.Cancelled, order.Rejected]:
            self.order = None

        print('-' * 50, 'ORDER BEGIN notify_order', datetime.now())
        print(order)
        print('-' * 50, 'ORDER END')

    def notify_trade(self, trade):
        print('-' * 50, 'TRADE BEGIN notify_trade', datetime.now())
        print(trade)
        print('-' * 50, 'TRADE END')

    def notify_data(self, data, status, *args, **kwargs):
        dn = data._name
        dt = datetime.now()
        msg = f'Data Status: {data._getstatusname(status)}'
        print(str(dt) + dn + msg)
        if data._getstatusname(status) == 'LIVE':
            self.live_data = True
            self.datastatus = 1
        else:
            self.live_data = False

def main():
    # run flash server
    thread = StoppableThread(target=api_server_start, args=(API_PORT, API_REV))
    thread.start()
    thread.join()

    # host = "localhost"
    host = "192.168.0.108"
    store = MTraderStore(host=host, debug=False, datatimeout=10)

    # comment next 2 lines to use backbroker for backtesting with MTraderStore
    cerebro = bt.Cerebro()
    
    cerebro.addstrategy(SmaCross, store)
    broker = store.getbroker(use_positions=True)
    cerebro.setbroker(broker)

    start_date = datetime.now() - timedelta(hours=10)

    data = store.getdata(
        dataname="BTCUSD", 
        timeframe=bt.TimeFrame.Minutes,
        compression=1,
        fromdate=start_date, 
        # historical=True,
        # useask=True,
    )
    
    # cerebro.resampledata(data,
    #                  timeframe=bt.TimeFrame.Minutes,
    #                  compression=5
    #                  )

    # data = store.getdata(dataname="XAUUSD.c", timeframe=bt.TimeFrame.Ticks, fromdate=start_date) #, useask=True, historical=True)
    #                 the parameter "useask" will request the ask price insetad if the default bid price

    cerebro.adddata(data)
    # cerebro.adddata(data2)

    cerebro.run(stdstats=False)
    # cerebro.plot(style='candlestick', volume=False)

if __name__ == "__main__":
    main()